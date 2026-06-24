"""SQL query layer — full implementation (replaces U1 stub).

All public functions return DataFrames cached with @cache.memoize.
Cache is initialized in run.py via init_cache(server).

Patterns followed from competitive-shelf-intelligence/app/data.py:
- get_conn() imported inside function body (no import-time DB connections)
- @cache.memoize with per-function TTL
- every except Exception logs before returning empty DataFrame
- LEFT JOIN from product_master spine (50 SKUs) with row-count assert

Unit note:
- scan_data: units/store/week
- sku_inventory / production_schedule: cases
- capacity functions need consistent units — get_sop_summary converts cases
  to units using case_pack_qty before calling compute_stockout_date
"""

from __future__ import annotations

import logging
import os
from datetime import date, timedelta

import pandas as pd
from flask_caching import Cache

from app.analytics.capacity import (
    compute_decision_deadline,
    compute_stockout_date,
    detect_shared_line_conflicts,
)
from app.analytics.forecast import build_rolling_forecast
from app.analytics.oos_correction import correct_velocity, detect_oos_periods

cache = Cache()
logger = logging.getLogger(__name__)

_DEMO_AS_OF_DATE = "2025-11-01"   # reference date for demo stockout story
_SPINE_SKU_COUNT = 50
_SCENARIO_CACHE_TTL = 60          # seconds — scenario-keyed cache
_DEMAND_CACHE_TTL   = 3600        # 1 hour — stable demand data

# Bound the scan-data history window at the SQL level so the all-SKU path
# never loads the full table into pandas (the OOM cause). The forecast needs
# >= 52 weeks for STL(period=52); 78 weeks gives that plus ~1.5 years for the
# OOS seasonal index, while cutting the load from ~1.34M rows to ~0.75M.
_HISTORY_WEEKS = 78

_LIVE_COMPUTE = os.environ.get("LIVE_COMPUTE", "").lower() in ("1", "true", "yes")


def init_cache(server) -> None:
    cache_dir = os.environ.get("CACHE_DIR", "/cache")
    try:
        cache.init_app(server, config={
            "CACHE_TYPE": "FileSystemCache",
            "CACHE_DIR": cache_dir,
            "CACHE_DEFAULT_TIMEOUT": _DEMAND_CACHE_TTL,
        })
    except Exception:
        cache.init_app(server, config={
            "CACHE_TYPE": "SimpleCache",
            "CACHE_DEFAULT_TIMEOUT": _DEMAND_CACHE_TTL,
        })


# ---------------------------------------------------------------------------
# Scenario parameter validation
# ---------------------------------------------------------------------------

def _clamp_scenario(
    promo_lift_pct: float,
    new_doors: int,
    lead_time_slip_weeks: int,
) -> tuple[float, int, int]:
    """Clamp scenario params to valid ranges, log warnings for out-of-range inputs."""
    original = (promo_lift_pct, new_doors, lead_time_slip_weeks)
    promo_lift_pct    = max(0.0, min(1.0, promo_lift_pct))
    new_doors         = max(0, min(5000, new_doors))
    lead_time_slip_weeks = max(0, min(12, lead_time_slip_weeks))
    clamped = (promo_lift_pct, new_doors, lead_time_slip_weeks)
    if clamped != original:
        logger.warning("Scenario params clamped: original=%s → clamped=%s", original, clamped)
    return clamped


# ---------------------------------------------------------------------------
# Raw data queries
# ---------------------------------------------------------------------------

@cache.memoize(timeout=_DEMAND_CACHE_TTL)
def get_product_master() -> pd.DataFrame:
    """All 50 Cinderhaven SKUs from product_master."""
    from app.db import get_conn
    try:
        with get_conn() as conn:
            df = pd.read_sql(
                """
                SELECT sku, product_name, product_line, case_pack_qty
                FROM product_master
                ORDER BY sku
                """,
                conn,
            )
            if len(df) != _SPINE_SKU_COUNT:
                logger.warning("get_product_master: expected %d rows, got %d",
                               _SPINE_SKU_COUNT, len(df))
            return df
    except Exception:
        logger.exception("get_product_master failed")
        return pd.DataFrame()


@cache.memoize(timeout=_DEMAND_CACHE_TTL)
def get_scan_data(sku: str | None = None) -> pd.DataFrame:
    """Scan data with is_authorized (from distribution_log) and is_promo flags.

    LEFT JOIN from scan_data spine. Aggregate channels are included but
    flagged (store_id ends with -AGG). OOS correction excludes them.

    History is bounded to the trailing _HISTORY_WEEKS window at the SQL level
    (both the all-SKU and single-SKU paths). This keeps the all-SKU forecast
    path from loading the full ~1.34M-row table into pandas — the OOM cause —
    while still returning the >= 52 weeks STL needs plus headroom for the OOS
    seasonal index. `dollars_sold` is intentionally not selected: no analytics
    or view consumes it.
    """
    from app.db import get_conn
    # CTE pre-computes promo flags on distinct (sku, week_ending) pairs (~3,900)
    # instead of a correlated EXISTS on every scan_data row (~840K). No f-string
    # interpolation — all dynamic values are bound parameters.
    _SQL_ALL = """
        WITH promo_flag AS (
            SELECT DISTINCT sub.sku, sub.week_ending
            FROM (
                SELECT DISTINCT sku, week_ending
                FROM scan_data
                WHERE week_ending <= %(as_of)s
                  AND week_ending >  %(history_start)s
            ) sub
            JOIN promotions p
                ON p.sku = sub.sku
                AND sub.week_ending BETWEEN p.start_week AND p.end_week
        )
        SELECT
            s.sku,
            s.store_id,
            s.week_ending,
            s.units_sold,
            CASE
                WHEN d.sku IS NOT NULL
                     AND d.authorized_date  <= s.week_ending
                     AND (d.deauthorized_date IS NULL
                          OR d.deauthorized_date > s.week_ending)
                THEN TRUE
                ELSE FALSE
            END                                             AS is_authorized,
            CASE WHEN pf.sku IS NOT NULL THEN TRUE
                 ELSE FALSE
            END                                             AS is_promo
        FROM scan_data s
        LEFT JOIN distribution_log d
            ON d.sku = s.sku AND d.store_id = s.store_id
        LEFT JOIN promo_flag pf
            ON pf.sku = s.sku AND pf.week_ending = s.week_ending
        WHERE s.week_ending <= %(as_of)s
          AND s.week_ending >  %(history_start)s
        ORDER BY s.sku, s.store_id, s.week_ending
    """
    _SQL_SKU = """
        WITH promo_flag AS (
            SELECT DISTINCT sub.sku, sub.week_ending
            FROM (
                SELECT DISTINCT sku, week_ending
                FROM scan_data
                WHERE sku = %(sku)s
                  AND week_ending <= %(as_of)s
                  AND week_ending >  %(history_start)s
            ) sub
            JOIN promotions p
                ON p.sku = sub.sku
                AND sub.week_ending BETWEEN p.start_week AND p.end_week
        )
        SELECT
            s.sku,
            s.store_id,
            s.week_ending,
            s.units_sold,
            CASE
                WHEN d.sku IS NOT NULL
                     AND d.authorized_date  <= s.week_ending
                     AND (d.deauthorized_date IS NULL
                          OR d.deauthorized_date > s.week_ending)
                THEN TRUE
                ELSE FALSE
            END                                             AS is_authorized,
            CASE WHEN pf.sku IS NOT NULL THEN TRUE
                 ELSE FALSE
            END                                             AS is_promo
        FROM scan_data s
        LEFT JOIN distribution_log d
            ON d.sku = s.sku AND d.store_id = s.store_id
        LEFT JOIN promo_flag pf
            ON pf.sku = s.sku AND pf.week_ending = s.week_ending
        WHERE s.sku = %(sku)s
          AND s.week_ending <= %(as_of)s
          AND s.week_ending >  %(history_start)s
        ORDER BY s.sku, s.store_id, s.week_ending
    """
    history_start = str(
        (pd.Timestamp(_DEMO_AS_OF_DATE) - pd.Timedelta(weeks=_HISTORY_WEEKS)).date()
    )
    try:
        if sku:
            sql = _SQL_SKU
            params = {"sku": sku, "as_of": _DEMO_AS_OF_DATE, "history_start": history_start}
        else:
            sql = _SQL_ALL
            params = {"as_of": _DEMO_AS_OF_DATE, "history_start": history_start}
        with get_conn() as conn:
            cur = conn.cursor()
            cur.execute(sql, params)
            cols = [desc[0] for desc in cur.description]
            return pd.DataFrame(cur.fetchall(), columns=cols)
    except Exception:
        logger.exception("get_scan_data failed")
        return pd.DataFrame()


@cache.memoize(timeout=_DEMAND_CACHE_TTL)
def get_true_demand(sku: str | None = None) -> pd.DataFrame:
    """OOS-corrected true_demand per (sku, store_id, week_ending).

    Calls detect_oos_periods() → correct_velocity() on the scan DataFrame.
    """
    scan = get_scan_data(sku=sku)
    if scan.empty:
        return pd.DataFrame()
    try:
        with_oos = detect_oos_periods(scan)
        corrected = correct_velocity(with_oos)
        return corrected
    except Exception:
        logger.exception("get_true_demand failed")
        return pd.DataFrame()


@cache.memoize(timeout=_DEMAND_CACHE_TTL)
def get_sku_inventory() -> pd.DataFrame:
    """Current on-hand inventory in cases and units (using case_pack_qty)."""
    from app.db import get_conn
    try:
        with get_conn() as conn:
            df = pd.read_sql(
                """
                SELECT
                    i.sku,
                    i.as_of_date,
                    i.on_hand_cases,
                    pm.case_pack_qty,
                    i.on_hand_cases * pm.case_pack_qty  AS on_hand_units
                FROM sku_inventory i
                JOIN product_master pm ON pm.sku = i.sku
                ORDER BY i.sku
                """,
                conn,
            )
            if len(df) != _SPINE_SKU_COUNT:
                logger.warning("get_sku_inventory: expected %d rows, got %d",
                               _SPINE_SKU_COUNT, len(df))
            return df
    except Exception:
        logger.exception("get_sku_inventory failed")
        return pd.DataFrame()


@cache.memoize(timeout=_DEMAND_CACHE_TTL)
def get_production_schedule() -> pd.DataFrame:
    """Booked production runs converted to units."""
    from app.db import get_conn
    try:
        with get_conn() as conn:
            return pd.read_sql(
                """
                SELECT
                    ps.run_id,
                    ps.sku,
                    ps.line_id,
                    ps.scheduled_week,
                    ps.quantity_cases,
                    pm.case_pack_qty,
                    ps.quantity_cases * pm.case_pack_qty  AS quantity_units,
                    ps.status
                FROM production_schedule ps
                JOIN product_master pm ON pm.sku = ps.sku
                WHERE ps.status = 'booked'
                ORDER BY ps.scheduled_week, ps.sku
                """,
                conn,
            )
    except Exception:
        logger.exception("get_production_schedule failed")
        return pd.DataFrame()


@cache.memoize(timeout=_DEMAND_CACHE_TTL)
def get_sku_config() -> pd.DataFrame:
    """Per-SKU production config with product metadata."""
    from app.db import get_conn
    try:
        with get_conn() as conn:
            df = pd.read_sql(
                """
                SELECT
                    c.sku,
                    c.line_id,
                    c.lead_time_weeks,
                    c.min_run_cases,
                    pm.product_name,
                    pm.product_line,
                    pm.case_pack_qty
                FROM sku_production_config c
                JOIN product_master pm ON pm.sku = c.sku
                ORDER BY c.sku
                """,
                conn,
            )
            if len(df) != _SPINE_SKU_COUNT:
                logger.warning("get_sku_config: expected %d rows, got %d",
                               _SPINE_SKU_COUNT, len(df))
            return df
    except Exception:
        logger.exception("get_sku_config failed")
        return pd.DataFrame()


# ---------------------------------------------------------------------------
# Snapshot readers (used when LIVE_COMPUTE is off — the default)
# ---------------------------------------------------------------------------

@cache.memoize(timeout=_DEMAND_CACHE_TTL)
def _read_forecast_snapshot() -> pd.DataFrame:
    from app.db import get_conn
    try:
        with get_conn() as conn:
            cur = conn.cursor()
            cur.execute("""
                SELECT sku, product_name, product_line, weekly_forecast_mean,
                       current_inventory, stockout_date, decision_deadline,
                       days_until_deadline, deadline_flag, lead_time_weeks,
                       shared_line_conflict, conflict_skus, median_store_velocity
                FROM copack.forecast_snapshot
                ORDER BY sku
            """)
            cols = [desc[0] for desc in cur.description]
            df = pd.DataFrame(cur.fetchall(), columns=cols)
        df["stockout_date"] = pd.to_datetime(df["stockout_date"])
        df["decision_deadline"] = pd.to_datetime(df["decision_deadline"])
        df["conflict_skus"] = df["conflict_skus"].apply(
            lambda x: x.split(",") if isinstance(x, str) and x else []
        )
        return df
    except Exception:
        logger.exception("_read_forecast_snapshot failed")
        return pd.DataFrame()


@cache.memoize(timeout=_DEMAND_CACHE_TTL)
def _read_doom_loop_snapshot(sku: str) -> pd.DataFrame:
    from app.db import get_conn
    try:
        with get_conn() as conn:
            cur = conn.cursor()
            cur.execute("""
                SELECT sku, week_ending, observed_units, corrected_units,
                       stores_dark, weekly_hidden_units, cumulative_hidden_units
                FROM copack.doom_loop_snapshot
                WHERE sku = %(sku)s
                ORDER BY week_ending
            """, {"sku": sku})
            cols = [desc[0] for desc in cur.description]
            df = pd.DataFrame(cur.fetchall(), columns=cols)
        if not df.empty:
            df["week_ending"] = pd.to_datetime(df["week_ending"])
        return df
    except Exception:
        logger.exception("_read_doom_loop_snapshot failed")
        return pd.DataFrame()


def _apply_scenario_to_snapshot(
    snap: pd.DataFrame,
    promo_lift_pct: float,
    new_doors: int,
    lead_time_slip_weeks: int,
) -> pd.DataFrame:
    """Adjust baseline snapshot values for scenario parameters.

    Uses linear approximations: stockout dates scale proportionally with
    the forecast change, decision deadlines shift by lead-time slip.
    """
    df = snap.copy()
    as_of = pd.Timestamp(_DEMO_AS_OF_DATE)
    baseline_forecast = df["weekly_forecast_mean"].copy()

    if promo_lift_pct > 0:
        df["weekly_forecast_mean"] *= (1 + promo_lift_pct)
    if new_doors > 0:
        df["weekly_forecast_mean"] += new_doors * df["median_store_velocity"].fillna(0)

    has_stockout = df["stockout_date"].notna()
    if has_stockout.any() and (promo_lift_pct > 0 or new_doors > 0):
        original_days = (
            df.loc[has_stockout, "stockout_date"] - as_of
        ).dt.days.astype(float)
        safe_forecast = df.loc[has_stockout, "weekly_forecast_mean"].replace(
            0, float("nan")
        )
        ratio = baseline_forecast[has_stockout] / safe_forecast
        adjusted_days = (original_days * ratio).round()
        df.loc[has_stockout, "stockout_date"] = (
            as_of + pd.to_timedelta(adjusted_days, unit="D")
        )

    lt = df["lead_time_weeks"].fillna(0).astype(float) + lead_time_slip_weeks
    df["decision_deadline"] = pd.NaT
    has_stockout = df["stockout_date"].notna()
    if has_stockout.any():
        df.loc[has_stockout, "decision_deadline"] = (
            df.loc[has_stockout, "stockout_date"]
            - pd.to_timedelta(lt[has_stockout] * 7, unit="D")
        )

    df["days_until_deadline"] = pd.array([pd.NA] * len(df), dtype="Int64")
    has_deadline = df["decision_deadline"].notna()
    if has_deadline.any():
        df.loc[has_deadline, "days_until_deadline"] = (
            (df.loc[has_deadline, "decision_deadline"] - as_of).dt.days
        )

    flags = []
    for days in df["days_until_deadline"]:
        if days is pd.NA or pd.isna(days):
            flags.append("OK")
        elif days < 0:
            flags.append("PAST_DUE")
        elif days < 14:
            flags.append("CRITICAL")
        elif days < 28:
            flags.append("WARNING")
        else:
            flags.append("OK")
    df["deadline_flag"] = flags

    return df


def _forecast_from_snapshot(
    promo_lift_pct: float = 0.0, new_doors: int = 0
) -> pd.DataFrame:
    """Synthesize per-week forecast from snapshot baseline (flat line per SKU)."""
    snap = _read_forecast_snapshot()
    if snap.empty:
        return pd.DataFrame()

    forecast_from = pd.Timestamp(_DEMO_AS_OF_DATE)
    rows = []
    for _, sku_row in snap.iterrows():
        fcst = float(sku_row["weekly_forecast_mean"])
        if promo_lift_pct > 0:
            fcst *= (1 + promo_lift_pct)
        if new_doors > 0:
            fcst += new_doors * float(sku_row.get("median_store_velocity", 0))

        for week_offset in range(1, 13):
            rows.append({
                "sku": sku_row["sku"],
                "week_ending": forecast_from + pd.Timedelta(weeks=week_offset),
                "forecast_units": fcst,
                "is_projected": True,
                "forecast_method": "snapshot",
            })

    return pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# Scenario-keyed analytics queries
# ---------------------------------------------------------------------------

@cache.memoize(timeout=_SCENARIO_CACHE_TTL)
def get_forecast(
    promo_lift_pct: float = 0.0,
    new_doors: int = 0,
    lead_time_slip_weeks: int = 0,
) -> pd.DataFrame:
    """12-week rolling demand forecast per SKU, with scenario parameters applied."""
    promo_lift_pct, new_doors, lead_time_slip_weeks = _clamp_scenario(
        promo_lift_pct, new_doors, lead_time_slip_weeks
    )
    if not _LIVE_COMPUTE:
        return _forecast_from_snapshot(promo_lift_pct, new_doors)
    true_demand = get_true_demand()
    if true_demand.empty:
        return pd.DataFrame()
    try:
        forecast_from = _get_forecast_from_week(true_demand)
        return build_rolling_forecast(
            true_demand,
            forecast_from_week=forecast_from,
            promo_lift_pct=promo_lift_pct,
            new_retailer_doors=new_doors,
        )
    except Exception:
        logger.exception("get_forecast failed")
        return pd.DataFrame()


@cache.memoize(timeout=_SCENARIO_CACHE_TTL)
def get_sop_summary(
    promo_lift_pct: float = 0.0,
    new_doors: int = 0,
    lead_time_slip_weeks: int = 0,
) -> pd.DataFrame:
    """Full S&OP summary: stockout date, decision deadline, conflict flags per SKU.

    Joins forecast → inventory → production schedule through the analytics
    pipeline. lead_time_slip_weeks adds N weeks to each SKU's lead_time_weeks
    for scenario modeling.
    """
    promo_lift_pct, new_doors, lead_time_slip_weeks = _clamp_scenario(
        promo_lift_pct, new_doors, lead_time_slip_weeks
    )
    if not _LIVE_COMPUTE:
        try:
            snap = _read_forecast_snapshot()
            if snap.empty:
                return pd.DataFrame()
            if promo_lift_pct > 0 or new_doors > 0 or lead_time_slip_weeks > 0:
                return _apply_scenario_to_snapshot(
                    snap, promo_lift_pct, new_doors, lead_time_slip_weeks
                )
            return snap
        except Exception:
            logger.exception("get_sop_summary (snapshot) failed")
            return pd.DataFrame()
    try:
        forecast = get_forecast(promo_lift_pct=promo_lift_pct, new_doors=new_doors)
        inventory = get_sku_inventory()
        schedule = get_production_schedule()
        sku_config = get_sku_config()

        if forecast.empty or inventory.empty or sku_config.empty:
            logger.warning("get_sop_summary: missing inputs — returning empty")
            return pd.DataFrame()

        # Convert inventory and schedule to units (already done by get_sku_inventory /
        # get_production_schedule using case_pack_qty in the query)
        inventory_for_cap = inventory[["sku", "on_hand_units"]].rename(
            columns={"on_hand_units": "on_hand_units"}
        )
        schedule_for_cap = (
            schedule[["sku", "line_id", "scheduled_week", "quantity_units", "status"]]
            if not schedule.empty else schedule
        )

        # Apply lead_time_slip to sku_config
        sku_config_adj = sku_config.copy()
        if lead_time_slip_weeks > 0:
            sku_config_adj["lead_time_weeks"] = (
                sku_config_adj["lead_time_weeks"] + lead_time_slip_weeks
            )

        # Step 1: stockout dates
        sop = compute_stockout_date(forecast, inventory_for_cap, schedule_for_cap)

        # Step 2: decision deadlines + flags
        # as_of_date anchors the demo to 2025-11-01; without it, today's date
        # makes every deadline PAST_DUE (all stockout dates are in 2025/early 2026).
        sop = compute_decision_deadline(
            sop, sku_config_adj, as_of_date=pd.Timestamp(_DEMO_AS_OF_DATE)
        )

        # Step 3: shared-line conflict detection
        sop = detect_shared_line_conflicts(sop, sku_config_adj)

        # Step 4: join product metadata
        meta = sku_config_adj[["sku", "product_name", "product_line"]].copy()
        sop = sop.merge(meta, on="sku", how="left")

        if len(sop) != _SPINE_SKU_COUNT:
            logger.warning("get_sop_summary: expected %d SKUs, got %d",
                           _SPINE_SKU_COUNT, len(sop))

        return sop
    except Exception:
        logger.exception("get_sop_summary failed")
        return pd.DataFrame()


# ---------------------------------------------------------------------------
# Doom loop data
# ---------------------------------------------------------------------------

def get_doom_loop_weekly(
    sku: str, period_weeks: int | None = None
) -> pd.DataFrame:
    """Weekly OOS/hidden-demand aggregates for a single SKU.

    Returns columns: week_ending, observed_units, corrected_units,
    stores_dark, weekly_hidden_units, cumulative_hidden_units.

    period_weeks: if set, filter to trailing N weeks from _DEMO_AS_OF_DATE.
    Cumulative hidden units are recomputed from the filtered window start.
    """
    if not _LIVE_COMPUTE:
        df = _read_doom_loop_snapshot(sku)
    else:
        td = get_true_demand(sku=sku)
        if td.empty or "is_oos" not in td.columns:
            return pd.DataFrame()
        td = td.copy()
        td["week_ending"] = pd.to_datetime(td["week_ending"])

        df = (
            td.groupby("week_ending")
            .agg(
                observed_units=("units_sold", "sum"),
                corrected_units=("true_demand", "sum"),
                stores_dark=("is_oos", "sum"),
            )
            .reset_index()
        )
        wh = (
            td[td["is_oos"]]
            .groupby("week_ending")["true_demand"]
            .sum()
            .rename("weekly_hidden_units")
            .reset_index()
        )
        df = df.merge(wh, on="week_ending", how="left")
        df["weekly_hidden_units"] = df["weekly_hidden_units"].fillna(0.0)
        df = df.sort_values("week_ending")
        df["cumulative_hidden_units"] = df["weekly_hidden_units"].cumsum()

    if df.empty:
        return df

    if period_weeks is not None:
        cutoff = pd.Timestamp(_DEMO_AS_OF_DATE) - pd.Timedelta(weeks=period_weeks)
        df = df[df["week_ending"] >= cutoff].copy()
        if not df.empty:
            df["cumulative_hidden_units"] = df["weekly_hidden_units"].cumsum()

    return df


# ---------------------------------------------------------------------------
# Export helpers
# ---------------------------------------------------------------------------

def export_sop_excel(
    sop_df: pd.DataFrame,
    scenario_params: dict | None = None,
) -> bytes:
    """Render the S&OP summary as a formatted .xlsx workbook.

    Returns raw bytes suitable for dcc.Download.

    Sheet 1 "Production Decision Brief": per-SKU rows, header navy fill,
        CRITICAL/PAST_DUE rows red, WARNING rows orange.
    Sheet 2 "Scenario Parameters": records the scenario inputs + generated_at.
    """
    import io
    from datetime import datetime

    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    from app.constants import CHICAGO, SURFACE_FAIL, SURFACE_WARN, TEAL

    def _hex(color: str) -> str:
        return color.lstrip("#")

    NAVY_FILL   = PatternFill("solid", fgColor=_hex(CHICAGO))
    RED_FILL    = PatternFill("solid", fgColor=_hex(SURFACE_FAIL))
    ORANGE_FILL = PatternFill("solid", fgColor=_hex(SURFACE_WARN))
    GREEN_FILL  = PatternFill("solid", fgColor=_hex("#e4f5f0"))
    WHITE_FONT  = Font(name="Calibri", bold=True, color="FFFFFF")
    BOLD_FONT   = Font(name="Calibri", bold=True)

    SHEET1_COLS = [
        ("SKU",               "sku"),
        ("Product Name",      "product_name"),
        ("Product Line",      "product_line"),
        ("Forecast/wk (units)", "weekly_forecast_mean"),
        ("Current Inv (units)", "current_inventory"),
        ("Stockout Week",     "stockout_date"),
        ("Decision Deadline", "decision_deadline"),
        ("Days Until Deadline", "days_until_deadline"),
        ("Action Flag",       "deadline_flag"),
        ("Shared Line Conflict", "shared_line_conflict"),
    ]

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Production Decision Brief"

    # Header row
    for col_idx, (header, _) in enumerate(SHEET1_COLS, start=1):
        cell = ws1.cell(row=1, column=col_idx, value=header)
        cell.fill = NAVY_FILL
        cell.font = WHITE_FONT
        ws1.column_dimensions[get_column_letter(col_idx)].width = max(len(header) + 4, 16)

    # Data rows
    for row_idx, (_, row) in enumerate(sop_df.iterrows(), start=2):
        flag = str(row.get("deadline_flag", "OK"))
        if flag in ("PAST_DUE", "CRITICAL"):
            row_fill = RED_FILL
        elif flag == "WARNING":
            row_fill = ORANGE_FILL
        else:
            row_fill = GREEN_FILL

        for col_idx, (_, field) in enumerate(SHEET1_COLS, start=1):
            val = row.get(field)
            if isinstance(val, pd.Timestamp):
                val = val.date() if pd.notna(val) else None
            elif isinstance(val, bool):
                val = "Yes" if val else ""
            elif val is None or (not isinstance(val, str) and pd.isna(val)):
                val = None
            if field == "weekly_forecast_mean" and isinstance(val, (int, float)):
                val = int(round(val))
            cell = ws1.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = row_fill
            if isinstance(val, date):
                cell.number_format = "YYYY-MM-DD"

    # Footer note
    note_row = len(sop_df) + 3
    ws1.cell(row=note_row, column=1,
             value="Generated by Lailara LLC Production Demand Forecast. "
                   "Data: Cinderhaven Provisions LLC (synthetic). "
                   f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M')}")

    # Sheet 2: Scenario Parameters
    ws2 = wb.create_sheet("Scenario Parameters")
    params = scenario_params or {}
    ws2.cell(row=1, column=1, value="Parameter").font = BOLD_FONT
    ws2.cell(row=1, column=2, value="Value").font = BOLD_FONT
    rows = [
        ("Promo Lift %", f"{params.get('promo_lift_pct', 0.0)*100:.0f}%"),
        ("New Retailer Doors", str(params.get("new_doors", 0))),
        ("Lead-Time Slip (weeks)", str(params.get("lead_time_slip_weeks", 0))),
        ("Generated At", datetime.now().strftime("%Y-%m-%d %H:%M")),
    ]
    for i, (k, v) in enumerate(rows, start=2):
        ws2.cell(row=i, column=1, value=k)
        ws2.cell(row=i, column=2, value=v)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def export_sop_pdf(
    sop_df: pd.DataFrame,
    scenario_params: dict | None = None,
    hero_fig=None,
) -> bytes:
    """Render the S&OP decision brief as a PDF via WeasyPrint + Jinja2.

    Returns raw PDF bytes.

    WeasyPrint requires native system libraries (pango, cairo, gdk-pixbuf).
    On Linux (Fly.io), these are installed in the Dockerfile.
    On Windows, this will raise ImportError — catch at the call site.

    hero_fig: optional Plotly go.Figure for the doom loop hero case.
    If kaleido is not installed, the SVG section is omitted gracefully.
    """
    import pathlib
    from datetime import datetime

    try:
        import weasyprint
        from jinja2 import Environment, FileSystemLoader
    except ImportError as e:
        raise ImportError(
            f"PDF export requires WeasyPrint and Jinja2 system libraries: {e}. "
            "This is only supported in the Fly.io container (not on Windows)."
        ) from e

    scenario = scenario_params or {}
    promo = scenario.get("promo_lift_pct", 0.0)
    doors = scenario.get("new_doors", 0)
    slip  = scenario.get("lead_time_slip_weeks", 0)

    parts = []
    if promo:
        parts.append(f"+{int(promo*100)}% promo lift")
    if doors:
        parts.append(f"{doors:,} new doors")
    if slip:
        parts.append(f"+{slip}wk lead-time slip")
    scenario_label = f"Scenario: {', '.join(parts)}" if parts else "Baseline"

    # KPI summary
    total_skus       = len(sop_df)
    need_action      = int(sop_df["deadline_flag"].isin(["PAST_DUE", "CRITICAL", "WARNING"]).sum())
    critical_conf    = int(
        sop_df["shared_line_conflict"].sum()
    ) if "shared_line_conflict" in sop_df.columns else 0

    # Build table rows for template
    sop_rows = []
    for _, row in sop_df.iterrows():
        flag = str(row.get("deadline_flag", "OK"))
        css_class = {
            "PAST_DUE": "past-due",
            "CRITICAL": "critical",
            "WARNING":  "warning",
            "OK":       "ok",
        }.get(flag, "ok")
        stockout = row.get("stockout_date")
        deadline = row.get("decision_deadline")
        days = row.get("days_until_deadline")
        sop_rows.append({
            "sku":               row.get("sku", ""),
            "product_name":      row.get("product_name", ""),
            "product_line":      row.get("product_line", ""),
            "weekly_forecast_mean": f"{row.get('weekly_forecast_mean', 0):.0f}",
            "stockout_label":    _fmt_date_pdf(stockout),
            "deadline_label":    _fmt_date_pdf(deadline),
            "days_left":         str(int(days)) if days is not None and pd.notna(days) else "—",
            "deadline_flag":     flag,
            "shared_line_conflict": bool(row.get("shared_line_conflict", False)),
            "css_class":         css_class,
        })

    # Optional: hero SVG from Plotly figure
    hero_svg = None
    if hero_fig is not None:
        try:
            hero_svg = hero_fig.to_image(format="svg").decode("utf-8")
        except Exception:
            logger.warning("export_sop_pdf: kaleido not available, skipping hero SVG")

    # Render Jinja2 template
    template_dir = pathlib.Path(__file__).parent / "templates"
    env = Environment(loader=FileSystemLoader(str(template_dir)), autoescape=True)
    tmpl = env.get_template("export_pdf.html")
    html_str = tmpl.render(
        generated_at=datetime.now().strftime("%Y-%m-%d %H:%M"),
        scenario_label=scenario_label,
        total_skus=total_skus,
        need_action=need_action,
        critical_conflicts=critical_conf,
        sop_rows=sop_rows,
        hero_svg=hero_svg,
    )

    return weasyprint.HTML(string=html_str).write_pdf()


def _fmt_date_pdf(d) -> str:
    try:
        if d is None or pd.isna(d):
            return "—"
        ts = pd.Timestamp(d)
        if pd.isna(ts):
            return "—"
        return str(ts.day) + ts.strftime(" %b %Y")
    except Exception:
        s = str(d) if d else ""
        return "—" if not s or s == "NaT" else s[:10]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _get_forecast_from_week(true_demand_df: pd.DataFrame) -> str:
    """Return the last week_ending in the dataset, or the demo reference date."""
    if true_demand_df.empty:
        return _DEMO_AS_OF_DATE
    try:
        last_week = pd.to_datetime(true_demand_df["week_ending"]).max()
        return str(last_week.date())
    except Exception:
        return _DEMO_AS_OF_DATE
